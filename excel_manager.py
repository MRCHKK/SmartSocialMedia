import pandas as pd
import datetime
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config_manager

logger = logging.getLogger(__name__)

# ── kolory ──────────────────────────────────────────────────────────────────
COLOR_YEAR_HEADER   = "1F3864"   # ciemny granat – nagłówek roku
COLOR_MONTH_HEADER  = "2E75B6"   # niebieski – nagłówek miesiąca
COLOR_LOCATION_NAME = "D6E4F0"   # jasny błękit – nazwa lokalizacji
COLOR_TABLE_HEADER  = "BDD7EE"   # nagłówek tabeli (5,4,3,2,1…)
COLOR_SUMA_ROW      = "DDEBF7"   # wiersz SUMA OPINII
COLOR_DZIAL_ROW     = "FFFFFF"   # wiersze działów
COLOR_SUMMARY_HEADER = "FFF2CC"  # nagłówek prawej sekcji
COLOR_SUMMARY_DATA   = "FFFDE7"  # dane prawej sekcji

FONT_WHITE  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_BOLD   = Font(name="Arial", bold=True, size=10)
FONT_NORMAL = Font(name="Arial", size=10)


# ── helpers ──────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _center(ws, cell_ref):
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


# ── obliczenia dla jednej lokalizacji w jednym miesiącu ──────────────────────

def _licz_gwiazdy(df_miesiac):
    """Zwraca dict {5:n, 4:n, 3:n, 2:n, 1:n} dla podanego slice'a."""
    counts = df_miesiac["Ocena"].value_counts()
    return {star: int(counts.get(star, 0)) for star in [5, 4, 3, 2, 1]}


def _srednia(gwiazdy):
    """Ważona średnia z rozkładu gwiazdek."""
    total = sum(gwiazdy.values())
    if total == 0:
        return None
    return round(sum(s * n for s, n in gwiazdy.items()) / total, 6)


def _buduj_dane_lokalizacji(df_lok, dzialy_kolejnosc, aktywne_kategorie):
    """
    Dla slice'a df jednej lokalizacji zwraca strukturę:
    {
        "gwiazdy_suma": {5:n,...},
        "dzialy": { "NAZWA DZIAŁU": {5:n,...}, ... }
    }
    Działy w kolejności z config + te które wystąpiły w danych i są aktywne.
    """
    # Kopia df_lok by bezpiecznie modyfikować Kategorie nieaktywne na Ogólne
    df_lok = df_lok.copy()
    if not df_lok.empty:
        df_lok.loc[~df_lok["Kategoria"].isin(aktywne_kategorie), "Kategoria"] = "Ogólne"

    gwiazdy_suma = _licz_gwiazdy(df_lok)

    dzialy_data = {}
    wszystkie_dzialy = list(dzialy_kolejnosc)
    # dodaj działy które są w danych i są aktywne w konfiguracji
    for d in df_lok["Kategoria"].unique():
        if d not in wszystkie_dzialy and d in aktywne_kategorie:
            wszystkie_dzialy.append(d)

    for dzial in wszystkie_dzialy:
        df_d = df_lok[df_lok["Kategoria"] == dzial]
        dzialy_data[dzial] = _licz_gwiazdy(df_d)

    return {"gwiazdy_suma": gwiazdy_suma, "dzialy": dzialy_data}


# ── zapis jednego bloku lokalizacji na arkusz ────────────────────────────────

def _zapisz_blok_lokalizacji(ws, start_row, lokalizacja_nazwa, miesiac_label,
                              dane, laczna_liczba, srednia_wizytwki):
    """
    Zapisuje blok dla jednej lokalizacji w jednym miesiącu.
    Zwraca numer wiersza po ostatnim zapisanym wierszu (już po pustej linii).

    Struktura bloku:
      [1] nazwa lokalizacji (B, szeroki)
      [2] nagłówek: miesiac_label | 5 | 4 | 3 | 2 | 1 | SUMA | ŚREDNIA | (sep) | Łączna… | Średnia ocen
      [3] SUMA OPINII + gwiazdy + suma + srednia      | laczna_liczba | srednia_wizytowki
      [4+] wiersze działów
      [blank]
    """
    COL_LABEL  = 2   # B
    COL_5      = 3   # C
    COL_4      = 4   # D
    COL_3      = 5   # E
    COL_2      = 6   # F
    COL_1      = 7   # G
    COL_SUMA   = 8   # H
    COL_SREDNIA = 9  # I
    # COL 10 = J – separator (pusty)
    COL_LACZNA  = 11  # K
    COL_SREDNIA_WIZ = 12  # L

    border = _thin_border()
    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left",   vertical="center")

    row = start_row

    # ── wiersz 1: nazwa lokalizacji ──────────────────────────────────────────
    _set_cell(ws, row, COL_LABEL, lokalizacja_nazwa,
              font=FONT_BOLD,
              fill=_fill(COLOR_LOCATION_NAME),
              alignment=align_left)
    row += 1

    # ── wiersz 2: nagłówek kolumn ────────────────────────────────────────────
    header_font = Font(name="Arial", bold=True, size=9)
    _set_cell(ws, row, COL_LABEL,    miesiac_label,  font=header_font, fill=_fill(COLOR_TABLE_HEADER), alignment=align_center, border=border)
    for col, val in [(COL_5, 5), (COL_4, 4), (COL_3, 3), (COL_2, 2), (COL_1, 1)]:
        _set_cell(ws, row, col, val, font=header_font, fill=_fill(COLOR_TABLE_HEADER), alignment=align_center, border=border)
    _set_cell(ws, row, COL_SUMA,    "SUMA",    font=header_font, fill=_fill(COLOR_TABLE_HEADER), alignment=align_center, border=border)
    _set_cell(ws, row, COL_SREDNIA, "ŚREDNIA", font=header_font, fill=_fill(COLOR_TABLE_HEADER), alignment=align_center, border=border)

    # prawa sekcja – tylko w nagłówku
    _set_cell(ws, row, COL_LACZNA,      "Łączna liczba opinii wizytówki",
              font=header_font, fill=_fill(COLOR_SUMMARY_HEADER), alignment=align_center, border=border)
    _set_cell(ws, row, COL_SREDNIA_WIZ, "Średnia ocen",
              font=header_font, fill=_fill(COLOR_SUMMARY_HEADER), alignment=align_center, border=border)
    row += 1

    # ── wiersz 3: SUMA OPINII ────────────────────────────────────────────────
    suma_font = Font(name="Arial", bold=True, size=10)
    gw = dane["gwiazdy_suma"]
    total = sum(gw.values())
    srednia = _srednia(gw)

    _set_cell(ws, row, COL_LABEL, "SUMA OPINII", font=suma_font, fill=_fill(COLOR_SUMA_ROW), alignment=align_left, border=border)
    for col, star in [(COL_5, 5), (COL_4, 4), (COL_3, 3), (COL_2, 2), (COL_1, 1)]:
        v = gw[star] if gw[star] > 0 else 0
        _set_cell(ws, row, col, v if v > 0 else 0, font=FONT_NORMAL, fill=_fill(COLOR_SUMA_ROW), alignment=align_center, border=border)
    _set_cell(ws, row, COL_SUMA,    total,   font=suma_font, fill=_fill(COLOR_SUMA_ROW), alignment=align_center, border=border)
    _set_cell(ws, row, COL_SREDNIA, round(srednia, 6) if srednia else None,
              font=suma_font, fill=_fill(COLOR_SUMA_ROW), alignment=align_center, border=border)

    # placeholder lub dane z zewnątrz
    _set_cell(ws, row, COL_LACZNA,      laczna_liczba,
              font=suma_font, fill=_fill(COLOR_SUMMARY_DATA), alignment=align_center, border=border)
    _set_cell(ws, row, COL_SREDNIA_WIZ, srednia_wizytwki,
              font=suma_font, fill=_fill(COLOR_SUMMARY_DATA), alignment=align_center, border=border)
    row += 1

    # ── wiersze działów ──────────────────────────────────────────────────────
    for dzial_nazwa, dzial_gw in dane["dzialy"].items():
        d_total = sum(dzial_gw.values())
        d_srednia = _srednia(dzial_gw)

        _set_cell(ws, row, COL_LABEL, dzial_nazwa, font=FONT_NORMAL, fill=_fill(COLOR_DZIAL_ROW), alignment=align_left, border=border)
        for col, star in [(COL_5, 5), (COL_4, 4), (COL_3, 3), (COL_2, 2), (COL_1, 1)]:
            v = dzial_gw[star]
            _set_cell(ws, row, col, v if v > 0 else None, font=FONT_NORMAL, fill=_fill(COLOR_DZIAL_ROW), alignment=align_center, border=border)
        _set_cell(ws, row, COL_SUMA,    d_total,  font=FONT_NORMAL, fill=_fill(COLOR_DZIAL_ROW), alignment=align_center, border=border)
        _set_cell(ws, row, COL_SREDNIA, round(d_srednia, 6) if d_srednia else None,
                  font=FONT_NORMAL, fill=_fill(COLOR_DZIAL_ROW), alignment=align_center, border=border)
        row += 1

    # ── pusta linia separatora ───────────────────────────────────────────────
    row += 1
    return row


# ── nagłówki roku i miesiąca ─────────────────────────────────────────────────

def _zapisz_naglowek_roku(ws, row, rok):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = _fill(COLOR_YEAR_HEADER)
    ws.cell(row=row, column=1, value=rok).font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 2  # rok + pusta linia


def _zapisz_naglowek_miesiaca(ws, row, miesiac_nazwa):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = _fill(COLOR_MONTH_HEADER)
    ws.cell(row=row, column=1, value=miesiac_nazwa).font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 2  # miesiąc + pusta linia


# ── ustawienia arkusza ───────────────────────────────────────────────────────

def _ustaw_wymiary_kolumn(ws):
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 6   # 5★
    ws.column_dimensions["D"].width = 6   # 4★
    ws.column_dimensions["E"].width = 6   # 3★
    ws.column_dimensions["F"].width = 6   # 2★
    ws.column_dimensions["G"].width = 6   # 1★
    ws.column_dimensions["H"].width = 8   # SUMA
    ws.column_dimensions["I"].width = 11  # ŚREDNIA
    ws.column_dimensions["J"].width = 3   # separator
    ws.column_dimensions["K"].width = 32  # Łączna liczba opinii wizytówki
    ws.column_dimensions["L"].width = 13  # Średnia ocen


# ── główna funkcja ────────────────────────────────────────────────────────────

# Mapowanie numeru miesiąca na polską nazwę
MIESIACE = {
    1: "STYCZEŃ", 2: "LUTY", 3: "MARZEC", 4: "KWIECIEŃ",
    5: "MAJ", 6: "CZERWIEC", 7: "LIPIEC", 8: "SIERPIEŃ",
    9: "WRZESIEŃ", 10: "PAŹDZIERNIK", 11: "LISTOPAD", 12: "GRUDZIEŃ"
}


def generuj_zestawienie(date_range=None, wizytowki_dane=None):
    """
    Generuje plik Excel z zestawieniem opinii w stylu OPINIE_ZESTAWIENIE.

    Parameters
    ----------
    date_range : tuple[str, str] | None
        Para dat (od, do) w formacie "DD.MM.YYYY". None = wszystkie.
    wizytowki_dane : dict | None
        Słownik {nazwa_lokalizacji: {"laczna_liczba": int, "srednia": float}}
        z danymi pobranymi z zewnętrznego API (np. Google Maps).
        Jeśli None lub lokalizacja nie ma wpisu, użyte zostaną placeholdery.
    """
    config = config_manager.load_config()

    if not os.path.exists(config["CSV_DATABASE"]):
        raise FileNotFoundError(f"Brak pliku {config['CSV_DATABASE']}. Najpierw pobierz dane z API.")

    df_full = pd.read_csv(config["CSV_DATABASE"])
    df_full["Data_dt"] = pd.to_datetime(df_full["Data"], format="%d.%m.%Y %H:%M", errors="coerce")
    df = df_full.copy()

    if date_range is not None:
        try:
            d_from = pd.to_datetime(date_range[0], format="%d.%m.%Y")
            d_to   = pd.to_datetime(date_range[1], format="%d.%m.%Y") + pd.Timedelta(days=1, seconds=-1)
            df = df[(df["Data_dt"] >= d_from) & (df["Data_dt"] <= d_to)]
        except Exception as e:
            logger.error(f"Błąd parsowania dat: {e}")
            raise ValueError("Nieprawidłowy format daty. Użyj DD.MM.YYYY")

    if df.empty:
        raise ValueError("Brak opinii w wybranym przedziale czasowym.")

    # kolejność działów z config
    dzialy_config = config.get("DZIALY", [])
    dzialy_kolejnosc = list(dzialy_config.keys()) if isinstance(dzialy_config, dict) else list(dzialy_config)

    # Aktywne kategorie (by nie wyświetlać usuniętych pracowników/działów w osobnym wierszu)
    pracownicy_config = config.get("PRACOWNICY", [])
    pracownicy_lista = list(pracownicy_config.keys()) if isinstance(pracownicy_config, dict) else list(pracownicy_config)
    aktywne_kategorie = set(pracownicy_lista + dzialy_kolejnosc + ["Ogólne"])

    wizytowki_dane = wizytowki_dane or {}

    # ── grupowanie: rok → miesiąc → lokalizacja ──────────────────────────────
    df["rok"]     = df["Data_dt"].dt.year
    df["miesiac"] = df["Data_dt"].dt.month

    # lokalizacje z config (zachowana kolejność)
    lokalizacje_config = config.get("LOKALIZACJE", [])
    lokalizacje_kolejnosc = (
        list(lokalizacje_config.values()) if isinstance(lokalizacje_config, dict)
        else list(lokalizacje_config)
    )

    # jeśli config nie ma LOKALIZACJE, użyj unikalnych wartości z danych
    if not lokalizacje_kolejnosc:
        lokalizacje_kolejnosc = sorted(df["Lokalizacja"].unique().tolist())

    # Wstępne pogrupowanie dla optymalizacji (zapobiega powtarzającemu się filtrowaniu całego df_full)
    df_by_lok = {
        lok: group.copy()
        for lok, group in df_full.groupby("Lokalizacja")
    }

    # ── budowanie pliku ───────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Zestawienie"
    ws.sheet_view.showGridLines = False
    _ustaw_wymiary_kolumn(ws)

    current_row = 1

    for rok in sorted(df["rok"].unique()):
        df_rok = df[df["rok"] == rok]

        current_row = _zapisz_naglowek_roku(ws, current_row, int(rok))

        for miesiac in sorted(df_rok["miesiac"].unique()):
            df_mies = df_rok[df_rok["miesiac"] == miesiac]
            miesiac_label = f"{miesiac:02d}.{rok}"
            miesiac_nazwa = f"{MIESIACE.get(miesiac, str(miesiac))} {rok}"

            current_row = _zapisz_naglowek_miesiaca(ws, current_row, miesiac_nazwa)

            # Oblicz koniec bieżącego miesiąca dla skumulowanych statystyk
            if miesiac == 12:
                nast_rok = rok + 1
                nast_mies = 1
            else:
                nast_rok = rok
                nast_mies = miesiac + 1
            limit_daty = pd.to_datetime(f"01.{nast_mies:02d}.{nast_rok}", format="%d.%m.%Y")

            # Oblicz statystyki dla każdej lokalizacji do tego miesiąca włącznie (na podstawie pełnej bazy)
            statystyki_do_miesiaca = {}
            for lok in df_full["Lokalizacja"].dropna().unique():
                df_lok_all = df_by_lok.get(lok)
                if df_lok_all is not None and not df_lok_all.empty:
                    df_lok_do_mies = df_lok_all[df_lok_all["Data_dt"] < limit_daty]
                    count_do_mies = len(df_lok_do_mies)
                    avg_do_mies = round(df_lok_do_mies["Ocena"].mean(), 2) if count_do_mies > 0 else "—"
                else:
                    count_do_mies = 0
                    avg_do_mies = "—"
                
                statystyki_do_miesiaca[lok] = {
                    "laczna_liczba": count_do_mies,
                    "srednia": avg_do_mies
                }

            for lokalizacja in lokalizacje_kolejnosc:
                df_lok = df_mies[df_mies["Lokalizacja"] == lokalizacja]

                # pomiń jeśli brak danych (opcjonalne – można też pokazywać pusty blok)
                # W obecnym pliku wszystkie lokalizacje są zawsze pokazywane:
                dane = _buduj_dane_lokalizacji(df_lok, dzialy_kolejnosc, aktywne_kategorie)

                # priorytet: dane przekazane z zewnątrz, potem skumulowane z bazy CSV do danego miesiąca, na końcu placeholder
                info = wizytowki_dane.get(lokalizacja) or statystyki_do_miesiaca.get(lokalizacja, {})
                laczna_liczba    = info.get("laczna_liczba", "—")
                srednia_wizytowki = info.get("srednia", "—")

                current_row = _zapisz_blok_lokalizacji(
                    ws, current_row,
                    lokalizacja, miesiac_label,
                    dane,
                    laczna_liczba, srednia_wizytowki
                )

    # ── zapis pliku ───────────────────────────────────────────────────────────
    dzisiaj = datetime.datetime.now().strftime("%d.%m.%Y_%H-%M-%S")
    bazowa  = config["EXCEL_FILE"].replace(".xlsx", "")
    sciezka = f"{bazowa}_zestawienie_{dzisiaj}.xlsx"

    os.makedirs(os.path.dirname(sciezka), exist_ok=True) if os.path.dirname(sciezka) else None
    wb.save(sciezka)

    logger.info(f"Wygenerowano zestawienie: {sciezka}")
    print(f"Gotowe! Plik zestawienia: {sciezka}")
    return sciezka


def przenies_do_excel(wizualizuj=False, date_range=None):
    """Kompatybilność wsteczna z poprzednią wersją managera."""
    return generuj_zestawienie(date_range=date_range)